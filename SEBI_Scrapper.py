from urllib.request import urlopen
from bs4 import BeautifulSoup
import ssl
import openpyxl

# Ignore SSL certificate errors for https
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

wb = openpyxl.load_workbook("PDF_Database.xlsx")
sheet = wb.active

url = "https://www.sebi.gov.in/sebiweb/home/HomeAction.do?doListing=yes&sid=3&s"
html = urlopen(url,context=ctx).read()
soup = BeautifulSoup(html, "html.parser")
title_tags = soup.select("tbody a")

rowcounter = 1
for i in range(25):
    new_url = title_tags[i]["href"]
    html2 = urlopen(new_url,context=ctx).read()
    soup2 = BeautifulSoup(html2, "html.parser")
    pdf_tags = soup2.select("tbody a")
    
    if pdf_tags == []:
        pdf_frame = soup2.select("body iframe")
        link = pdf_frame[0]["src"].split('=')
        sheet.cell(row=rowcounter, column=1).value = title_tags[i].string
        sheet.cell(row=rowcounter, column=2).value = link[1]
        
    else:
        sheet.cell(row=rowcounter,column=1).value = title_tags[i].string
        for j in pdf_tags:
            if j.string != "":
                sheet.cell(row=rowcounter, column=2).value = j.string
                sheet.cell(row=rowcounter, column=3).value = j["href"]
            rowcounter+=1
    rowcounter+=1

wb.save("PDF_Database.xlsx")